import csv
import subprocess
import sys
import unicodedata
from datetime import datetime
from pathlib import Path


class FondoAcciones:
    ANIO_POR_DEFECTO = 2026
    FORMATO_INTERNO_FECHA = "%Y-%m-%d %H:%M"
    MOVIMIENTOS_SEMILLA = [
        {
            "nombre_accion": "ECOPETROL",
            "tipo": "compra",
            "estado": "aprobada",
            "fecha": "2026-03-07 20:47",
            "cantidad": 58,
            "precio_unidad": 2425,
            "comision": 14875,
            "precio_total": 155525,
        },
        {
            "nombre_accion": "ECOPETROL",
            "tipo": "compra",
            "estado": "aprobada",
            "fecha": "2026-03-17 09:46",
            "cantidad": 14,
            "precio_unidad": 2675,
            "comision": 14875,
            "precio_total": 52325,
        },
        {
            "nombre_accion": "ECOPETROL",
            "tipo": "compra",
            "estado": "aprobada",
            "fecha": "2026-03-14 18:46",
            "cantidad": 85,
            "precio_unidad": 2535,
            "comision": 14875,
            "precio_total": 230350,
        },
        {
            "nombre_accion": "ECOPETROL",
            "tipo": "compra",
            "estado": "aprobada",
            "fecha": "2026-03-23 09:17",
            "cantidad": 2100,
            "precio_unidad": 2730,
            "comision": 17055.67,
            "precio_total": 5750055.67,
        },
        {
            "nombre_accion": "ECOPETROL",
            "tipo": "compra",
            "estado": "aprobada",
            "fecha": "2026-04-16 13:31",
            "cantidad": 200,
            "precio_unidad": 2480,
            "comision": 7437.5,
            "precio_total": 503437.5,
        },
        {
            "nombre_accion": "ECOPETROL",
            "tipo": "venta",
            "estado": "aprobada",
            "fecha": "2026-04-14 10:22",
            "cantidad": 800,
            "precio_unidad": 2445,
            "comision": 7437.5,
            "precio_total": 1948562.5,
        },
        {
            "nombre_accion": "ECOPETROL",
            "tipo": "venta",
            "estado": "aprobada",
            "fecha": "2026-04-14 15:04",
            "cantidad": 757,
            "precio_unidad": 2415,
            "comision": 7437.5,
            "precio_total": 1820717.5,
        },
        {
            "nombre_accion": "MINEROS",
            "tipo": "compra",
            "estado": "aprobada",
            "fecha": "2026-04-14 10:23",
            "cantidad": 123,
            "precio_unidad": 14620,
            "comision": 7437.5,
            "precio_total": 1805697.5,
        },
        {
            "nombre_accion": "NUCO",
            "tipo": "compra",
            "estado": "aprobada",
            "fecha": "2026-03-24 08:47",
            "cantidad": 4,
            "precio_unidad": 53740,
            "comision": 14875,
            "precio_total": 229835,
        },
        {
            "nombre_accion": "PFCIBEST",
            "tipo": "compra",
            "estado": "aprobada",
            "id": "11157664276394",
            "fecha": "2026-03-10 12:43",
            "cantidad": 3,
            "precio_unidad": 64980,
            "comision": 14875,
            "precio_total": 209815,
        },
        {
            "nombre_accion": "PFCIBEST",
            "tipo": "compra",
            "estado": "aprobada",
            "fecha": "2026-03-17 09:45",
            "cantidad": 3,
            "precio_unidad": 63640,
            "comision": 14875,
            "precio_total": 205795,
        },
        {
            "nombre_accion": "PFCIBEST",
            "tipo": "compra",
            "estado": "aprobada",
            "fecha": "2026-04-18 00:00",
            "cantidad": 20,
            "precio_unidad": 67600,
            "comision": 7437,
            "precio_total": 1359437,
        },
        {
            "nombre_accion": "PFCIBEST",
            "tipo": "dividendo",
            "estado": "aprobada",
            "fecha": "2026-04-01 00:00",
            "cantidad": 6,
            "precio_unidad": 1128,
            "comision": 238,
            "precio_total": 6530,
        },
        {
            "nombre_accion": "MINEROS",
            "tipo": "dividendo",
            "estado": "aprobada",
            "fecha": "2026-04-27 00:00",
            "cantidad": 123,
            "precio_unidad": 89.01544715447154,
            "comision": 238,
            "precio_total": 10710.9,
        },
        {
            "nombre_accion": "ECOPETROL",
            "tipo": "dividendo",
            "estado": "aprobada",
            "fecha": "2026-04-30 00:00",
            "cantidad": 900,
            "precio_unidad": 121,
            "comision": 1324.91,
            "precio_total": 107575.09,
        },
        {
            "nombre_accion": "IBITCO",
            "tipo": "compra",
            "estado": "aprobada",
            "id": "11157664638327",
            "fecha": "2026-05-09 07:41",
            "cantidad": 1,
            "precio_unidad": 174960,
            "comision": 7437.5,
            "precio_total": 182397.5,
        },
        {
            "nombre_accion": "PARAUCOCO",
            "tipo": "compra",
            "estado": "cancelada",
            "fecha": "2026-04-14 10:36",
            "cantidad": 8,
            "precio_unidad": 17200,
            "comision": 0,
            "precio_total": 0,
        },
        {
            "nombre_accion": "PFCIBEST",
            "tipo": "compra",
            "estado": "cancelada",
            "fecha": "2026-03-14 18:47",
            "cantidad": 4,
            "precio_unidad": 61140,
            "comision": 0,
            "precio_total": 0,
        },
    ]
    MESES = {
        1: "ene",
        2: "feb",
        3: "mar",
        4: "abr",
        5: "may",
        6: "jun",
        7: "jul",
        8: "ago",
        9: "sep",
        10: "oct",
        11: "nov",
        12: "dic",
    }
    MESES_INVERSOS = {valor: clave for clave, valor in MESES.items()}
    ALIASES_COLUMNAS_IMPORTACION = {
        "nombre_accion": {"accion", "accin", "nombre accion", "nombre de la accion"},
        "tipo": {"tipo", "tipo de movimiento"},
        "estado": {"estado"},
        "cantidad": {"cantidad", "cantidad de acciones", "acciones"},
        "precio_unidad": {"precio unidad", "precio por unidad", "precio", "valor unidad"},
        "comision": {"comision", "comisiones", "comisin", "valor comision"},
        "fecha": {"fecha"},
    }
    RETIRO_PERSONAL = 50000.0
    FONDOS_ACTUALES = 53664.0
    DISPONIBLE_PARA_INVERTIR = 67344.0
    VALOR_REAL_HOY = 6192000.0

    def __init__(self):
        self.movimientos = [
            self._normalizar_movimiento(movimiento.copy())
            for movimiento in self.MOVIMIENTOS_SEMILLA
        ]

    def _parsear_numero(self, texto):
        valor = texto.strip().replace(" ", "")
        if not valor:
            raise ValueError("Ingresa un numero.")

        if "," in valor and "." in valor:
            valor = valor.replace(".", "").replace(",", ".")
        elif "," in valor:
            valor = valor.replace(",", ".")
        elif valor.count(".") > 1:
            valor = valor.replace(".", "")
        elif valor.count(".") == 1:
            entero, decimales = valor.split(".")
            if entero.isdigit() and decimales.isdigit() and len(decimales) == 3:
                valor = entero + decimales

        return float(valor)

    def _pedir_entero(self, mensaje, valor_actual=None):
        while True:
            texto = input(mensaje).strip()

            if not texto and valor_actual is not None:
                return valor_actual

            try:
                numero = self._parsear_numero(texto)
            except ValueError:
                print("Ingresa un numero entero valido.")
                continue

            if not numero.is_integer():
                print("Ingresa un numero entero valido.")
                continue

            return int(numero)

    def _pedir_flotante(self, mensaje, valor_actual=None):
        while True:
            texto = input(mensaje).strip()

            if not texto and valor_actual is not None:
                return valor_actual

            try:
                return self._parsear_numero(texto)
            except ValueError:
                print("Ingresa un numero valido.")

    def _pedir_texto(self, mensaje, valor_actual=None, convertir_mayusculas=False):
        while True:
            texto = input(mensaje).strip()

            if not texto and valor_actual is not None:
                return valor_actual

            if texto:
                return texto.upper() if convertir_mayusculas else texto.lower()

            print("Este campo no puede quedar vacio.")

    def _pedir_tipo(self, mensaje, valor_actual=None, valor_defecto=None):
        opciones = {
            "compra": "compra",
            "venta": "venta",
            "dividendo": "dividendo",
            "dividendos": "dividendo",
        }

        while True:
            valor = input(mensaje).strip().lower()

            if not valor:
                if valor_actual is not None:
                    return valor_actual
                if valor_defecto is not None:
                    return valor_defecto

            if valor in opciones:
                return opciones[valor]

            print("El tipo debe ser 'compra', 'venta' o 'dividendo'.")

    def _pedir_estado(self, mensaje, valor_actual=None):
        equivalencias = {
            "aprobada": "aprobada",
            "cancelada": "cancelada",
            "rechazada": "cancelada",
        }

        while True:
            valor = input(mensaje).strip().lower()

            if not valor and valor_actual is not None:
                return valor_actual

            if valor in equivalencias:
                return equivalencias[valor]

            print("El estado debe ser 'aprobada' o 'cancelada'.")

    def _parsear_fecha(self, texto):
        valor = texto.strip().lower()
        formatos = [
            ("%Y-%m-%d %H:%M", False),
            ("%Y-%m-%d", True),
            ("%m-%d %H:%M", False),
            ("%m-%d", True),
        ]

        for formato, sin_hora in formatos:
            try:
                fecha = datetime.strptime(valor, formato)
                if formato.startswith("%m"):
                    fecha = fecha.replace(year=self.ANIO_POR_DEFECTO)
                if sin_hora:
                    fecha = fecha.replace(hour=0, minute=0)
                return fecha
            except ValueError:
                pass

        partes = valor.split()
        if len(partes) in {3, 4}:
            try:
                dia = int(partes[0])
                mes = self.MESES_INVERSOS[partes[1]]
                anio = int(partes[2])
                hora = "00:00" if len(partes) == 3 else partes[3]
                return datetime.strptime(
                    f"{anio}-{mes:02d}-{dia:02d} {hora}", "%Y-%m-%d %H:%M"
                )
            except (KeyError, ValueError):
                pass

        raise ValueError

    def _pedir_fecha(self, mensaje, valor_actual=None):
        while True:
            texto = input(mensaje).strip()

            if not texto and valor_actual is not None:
                return valor_actual

            if not texto:
                print("La fecha no puede quedar vacia.")
                continue

            try:
                fecha = self._parsear_fecha(texto)
            except ValueError:
                print(
                    "Fecha invalida. Usa YYYY-MM-DD HH:MM, YYYY-MM-DD, MM-DD HH:MM, MM-DD o 07 mar 2026 20:47."
                )
                continue

            return fecha.strftime(self.FORMATO_INTERNO_FECHA)

    def _fecha_a_datetime(self, fecha):
        try:
            return datetime.strptime(fecha, self.FORMATO_INTERNO_FECHA)
        except ValueError:
            return datetime.strptime(f"{fecha} 00:00", self.FORMATO_INTERNO_FECHA)

    def _formatear_fecha(self, fecha):
        fecha_dt = self._fecha_a_datetime(fecha)
        return (
            f"{fecha_dt.day:02d} {self.MESES[fecha_dt.month]} "
            f"{fecha_dt.year} {fecha_dt.hour:02d}:{fecha_dt.minute:02d}"
        )

    def _formatear_numero(self, valor):
        numero = float(valor)

        if numero.is_integer():
            texto = f"{int(numero):,}"
        else:
            texto = f"{numero:,.2f}".rstrip("0").rstrip(".")

        return texto.replace(",", "_").replace(".", ",").replace("_", ".")

    def _normalizar_clave_importacion(self, texto):
        valor = str(texto).strip().lower().replace("_", " ")
        valor = (
            unicodedata.normalize("NFKD", valor)
            .encode("ascii", "ignore")
            .decode("ascii")
        )
        valor = "".join(caracter for caracter in valor if caracter.isalnum() or caracter == " ")
        return " ".join(valor.split())

    def _normalizar_tipo(self, valor):
        tipo = self._normalizar_clave_importacion(valor)
        if tipo == "dividendos":
            tipo = "dividendo"
        if tipo not in {"compra", "venta", "dividendo"}:
            raise ValueError("El tipo debe ser compra, venta o dividendo.")
        return tipo

    def _normalizar_estado(self, valor):
        equivalencias = {
            "aprobada": "aprobada",
            "cancelada": "cancelada",
            "rechazada": "cancelada",
        }
        estado = self._normalizar_clave_importacion(valor)
        if estado not in equivalencias:
            raise ValueError("El estado debe ser aprobada o cancelada.")
        return equivalencias[estado]

    def _normalizar_fecha_valor(self, fecha):
        if isinstance(fecha, datetime):
            return fecha.strftime(self.FORMATO_INTERNO_FECHA)

        texto = str(fecha).strip()
        if not texto:
            raise ValueError("La fecha no puede quedar vacia.")

        for formato in (self.FORMATO_INTERNO_FECHA, "%Y-%m-%d"):
            try:
                fecha_dt = datetime.strptime(texto, formato)
                if formato == "%Y-%m-%d":
                    fecha_dt = fecha_dt.replace(hour=0, minute=0)
                return fecha_dt.strftime(self.FORMATO_INTERNO_FECHA)
            except ValueError:
                pass

        return self._parsear_fecha(texto).strftime(self.FORMATO_INTERNO_FECHA)

    def _calcular_inversion_bruta(self, cantidad, precio_unidad, estado):
        if estado == "cancelada":
            return 0.0
        return cantidad * precio_unidad

    def _calcular_total(self, cantidad, precio_unidad, comision, estado, tipo):
        inversion_bruta = self._calcular_inversion_bruta(cantidad, precio_unidad, estado)
        if estado == "cancelada":
            return 0.0
        if tipo in {"venta", "dividendo"}:
            return inversion_bruta - comision
        return inversion_bruta + comision

    def _normalizar_movimiento(self, movimiento):
        movimiento_normalizado = {
            "nombre_accion": str(movimiento["nombre_accion"]).strip().upper(),
            "tipo": self._normalizar_tipo(movimiento["tipo"]),
            "estado": self._normalizar_estado(movimiento["estado"]),
            "fecha": self._normalizar_fecha_valor(movimiento["fecha"]),
            "cantidad": int(round(float(movimiento["cantidad"]))),
            "precio_unidad": float(movimiento["precio_unidad"]),
            "comision": float(movimiento.get("comision", 0) or 0),
        }

        movimiento_normalizado["precio_total"] = self._calcular_total(
            movimiento_normalizado["cantidad"],
            movimiento_normalizado["precio_unidad"],
            movimiento_normalizado["comision"],
            movimiento_normalizado["estado"],
            movimiento_normalizado["tipo"],
        )

        if "id" in movimiento and movimiento["id"] not in (None, ""):
            movimiento_normalizado["id"] = str(movimiento["id"]).strip()

        return movimiento_normalizado

    def _valor_mostrable(self, texto):
        return texto.capitalize()

    def _tabla_markdown(self, filas, columnas_derecha=None):
        if not filas:
            print("No hay datos para mostrar.")
            return

        columnas_derecha = columnas_derecha or set()
        columnas = list(filas[0].keys())
        anchos = {}

        for columna in columnas:
            valor_mas_largo = max(len(str(fila[columna])) for fila in filas)
            anchos[columna] = max(len(columna), valor_mas_largo)

        encabezado = "| " + " | ".join(
            columna.ljust(anchos[columna]) for columna in columnas
        ) + " |"

        separador = "| " + " | ".join(
            ("-" * (anchos[columna] - 1) + ":")
            if columna in columnas_derecha
            else (":" + "-" * (anchos[columna] - 1))
            for columna in columnas
        ) + " |"

        print(encabezado)
        print(separador)

        for fila in filas:
            celdas = []
            for columna in columnas:
                valor = str(fila[columna])
                if columna in columnas_derecha:
                    celdas.append(valor.rjust(anchos[columna]))
                else:
                    celdas.append(valor.ljust(anchos[columna]))
            print("| " + " | ".join(celdas) + " |")

    def _filas_movimientos(self, movimientos):
        filas = []
        for movimiento in movimientos:
            filas.append(
                {
                    "Acción": movimiento["nombre_accion"],
                    "Tipo": self._valor_mostrable(movimiento["tipo"]),
                    "Estado": self._valor_mostrable(movimiento["estado"]),
                    "Cantidad": self._formatear_numero(movimiento["cantidad"]),
                    "Precio unidad": self._formatear_numero(
                        movimiento["precio_unidad"]
                    ),
                    "Comisión": self._formatear_numero(movimiento["comision"]),
                    "Total": self._formatear_numero(movimiento["precio_total"]),
                    "Fecha": self._formatear_fecha(movimiento["fecha"]),
                }
            )
        return filas

    def _mostrar_tabla_movimientos(self, movimientos, titulo):
        if not movimientos:
            print("No hay movimientos registrados.")
            return

        print(f"\n--- {titulo} ---")
        self._tabla_markdown(
            self._filas_movimientos(movimientos),
            columnas_derecha={"Cantidad", "Precio unidad", "Comisión", "Total"},
        )

    def _filas_movimientos(self, movimientos):
        filas = []
        for movimiento in movimientos:
            inversion_bruta = self._calcular_inversion_bruta(
                movimiento["cantidad"],
                movimiento["precio_unidad"],
                movimiento["estado"],
            )
            filas.append(
                {
                    "Accion": movimiento["nombre_accion"],
                    "Tipo": self._valor_mostrable(movimiento["tipo"]),
                    "Estado": self._valor_mostrable(movimiento["estado"]),
                    "Cantidad": self._formatear_numero(movimiento["cantidad"]),
                    "Precio unidad": self._formatear_numero(
                        movimiento["precio_unidad"]
                    ),
                    "Inversion": self._formatear_numero(inversion_bruta),
                    "Comision": self._formatear_numero(movimiento["comision"]),
                    "Total": self._formatear_numero(movimiento["precio_total"]),
                    "Fecha": self._formatear_fecha(movimiento["fecha"]),
                }
            )
        return filas

    def _mostrar_tabla_movimientos(self, movimientos, titulo):
        if not movimientos:
            print("No hay movimientos registrados.")
            return

        print(f"\n--- {titulo} ---")
        self._tabla_markdown(
            self._filas_movimientos(movimientos),
            columnas_derecha={
                "Cantidad",
                "Precio unidad",
                "Inversion",
                "Comision",
                "Total",
            },
        )

    def _seleccionar_movimiento(self, accion):
        if not self.movimientos:
            print(f"No hay movimientos para {accion}.")
            return None

        print(f"\n--- MOVIMIENTOS DISPONIBLES PARA {accion.upper()} ---")
        for indice, movimiento in enumerate(self.movimientos, start=1):
            print(
                f"{indice}. {movimiento['nombre_accion']} | "
                f"{self._valor_mostrable(movimiento['tipo'])} | "
                f"{self._valor_mostrable(movimiento['estado'])} | "
                f"{self._formatear_numero(movimiento['cantidad'])} | "
                f"{self._formatear_numero(movimiento['precio_total'])} | "
                f"{self._formatear_fecha(movimiento['fecha'])}"
            )

        while True:
            seleccion = input("Elige el numero del movimiento: ").strip()

            try:
                indice = int(seleccion)
            except ValueError:
                print("Ingresa un numero valido.")
                continue

            if 1 <= indice <= len(self.movimientos):
                return indice - 1

            print("Ese numero no corresponde a ningun movimiento.")

    def insertar_movimiento(self):
        print("\n--- INGRESAR NUEVO MOVIMIENTO ---")

        nombre_accion = self._pedir_texto(
            "Nombre de la accion: ", convertir_mayusculas=True
        )
        tipo = self._pedir_tipo(
            "Tipo de movimiento (compra/venta/dividendo) [Enter = compra]: ",
            valor_defecto="compra",
        )
        estado = self._pedir_estado("Estado (aprobada/cancelada): ")
        fecha = self._pedir_fecha(
            "Fecha (YYYY-MM-DD HH:MM, YYYY-MM-DD, MM-DD HH:MM o MM-DD): "
        )
        cantidad = self._pedir_entero("Cantidad de acciones: ")
        precio_unidad = self._pedir_flotante("Precio por unidad: ")
        comision = self._pedir_flotante("Comision (si no hay, escribe 0): ")
        precio_total = self._calcular_total(
            cantidad, precio_unidad, comision, estado, tipo
        )

        movimiento = {
            "nombre_accion": nombre_accion,
            "tipo": tipo,
            "estado": estado,
            "fecha": fecha,
            "cantidad": cantidad,
            "precio_unidad": precio_unidad,
            "comision": comision,
            "precio_total": precio_total,
        }

        self.movimientos.append(movimiento)
        print("Movimiento guardado correctamente.\n")

    def modificar_movimiento(self):
        indice = self._seleccionar_movimiento("modificar")
        if indice is None:
            return

        movimiento = self.movimientos[indice]

        print("\nDeja el campo vacio si quieres conservar el valor actual.")

        movimiento["nombre_accion"] = self._pedir_texto(
            f"Nombre de la accion [{movimiento['nombre_accion']}]: ",
            valor_actual=movimiento["nombre_accion"],
            convertir_mayusculas=True,
        )
        movimiento["tipo"] = self._pedir_tipo(
            f"Tipo de movimiento [{movimiento['tipo']}]: ",
            valor_actual=movimiento["tipo"],
        )
        movimiento["estado"] = self._pedir_estado(
            f"Estado [{movimiento['estado']}]: ",
            valor_actual=movimiento["estado"],
        )
        movimiento["fecha"] = self._pedir_fecha(
            f"Fecha [{self._formatear_fecha(movimiento['fecha'])}]: ",
            valor_actual=movimiento["fecha"],
        )
        movimiento["cantidad"] = self._pedir_entero(
            f"Cantidad de acciones [{self._formatear_numero(movimiento['cantidad'])}]: ",
            valor_actual=movimiento["cantidad"],
        )
        movimiento["precio_unidad"] = self._pedir_flotante(
            f"Precio por unidad [{self._formatear_numero(movimiento['precio_unidad'])}]: ",
            valor_actual=movimiento["precio_unidad"],
        )
        movimiento["comision"] = self._pedir_flotante(
            f"Comision [{self._formatear_numero(movimiento['comision'])}]: ",
            valor_actual=movimiento["comision"],
        )
        movimiento["precio_total"] = self._calcular_total(
            movimiento["cantidad"],
            movimiento["precio_unidad"],
            movimiento["comision"],
            movimiento["estado"],
            movimiento["tipo"],
        )

        print("Movimiento modificado correctamente.")

    def eliminar_movimiento(self):
        indice = self._seleccionar_movimiento("eliminar")
        if indice is None:
            return

        movimiento = self.movimientos[indice]
        confirmacion = input(
            f"Confirma eliminar {movimiento['nombre_accion']} del {self._formatear_fecha(movimiento['fecha'])} (s/n): "
        ).strip().lower()

        if confirmacion == "s":
            eliminado = self.movimientos.pop(indice)
            print(
                f"Movimiento eliminado: {eliminado['nombre_accion']} - {self._formatear_fecha(eliminado['fecha'])}."
            )
            return

        print("No se elimino ningun movimiento.")

    def _mostrar_grupo_movimientos(self, movimientos, titulo, mensaje_vacio):
        if not movimientos:
            print(f"\n--- {titulo} ---")
            print(mensaje_vacio)
            return

        self._mostrar_tabla_movimientos(movimientos, titulo)

    def _filas_resumen_movimientos_por_accion(self):
        resumen = {}

        for movimiento in self.movimientos:
            accion = movimiento["nombre_accion"]
            if accion not in resumen:
                resumen[accion] = {
                    "Accion": accion,
                    "Cantidad comprada": 0,
                    "Cantidad vendida": 0,
                    "Cantidad cancelada": 0,
                    "Total inversiones": 0.0,
                    "Total cancelado": 0.0,
                    "Ventas netas": 0.0,
                }

            if movimiento["estado"] == "cancelada":
                resumen[accion]["Cantidad cancelada"] += movimiento["cantidad"]
                resumen[accion]["Total cancelado"] += (
                    movimiento["cantidad"] * movimiento["precio_unidad"]
                )
                continue

            if movimiento["tipo"] == "compra":
                resumen[accion]["Cantidad comprada"] += movimiento["cantidad"]
                resumen[accion]["Total inversiones"] += movimiento["precio_total"]
            else:
                resumen[accion]["Cantidad vendida"] += movimiento["cantidad"]
                resumen[accion]["Ventas netas"] += movimiento["precio_total"]

        filas = []
        for fila in resumen.values():
            cantidad_acciones = fila["Cantidad comprada"] - fila["Cantidad vendida"]
            precio_compra_promedio = (
                fila["Total inversiones"] / fila["Cantidad comprada"]
                if fila["Cantidad comprada"] > 0
                else 0.0
            )
            ganancia_perdida = fila["Ventas netas"] - (
                fila["Cantidad vendida"] * precio_compra_promedio
            )

            filas.append(
                {
                    "Accion": fila["Accion"],
                    "Cantidad acciones": self._formatear_numero(cantidad_acciones),
                    "Total inversiones": self._formatear_numero(
                        fila["Total inversiones"]
                    ),
                    "Cantidad cancelada": self._formatear_numero(
                        fila["Cantidad cancelada"]
                    ),
                    "Total cancelado": self._formatear_numero(fila["Total cancelado"]),
                    "Cantidad vendida": self._formatear_numero(
                        fila["Cantidad vendida"]
                    ),
                    "Precio compra promedio": self._formatear_numero(
                        precio_compra_promedio
                    ),
                    "Ganancia / perdida": self._formatear_numero(ganancia_perdida),
                }
            )

        return filas

    def ver_movimientos(self):
        compras_aprobadas = [
            movimiento
            for movimiento in self.movimientos
            if movimiento["estado"] == "aprobada" and movimiento["tipo"] == "compra"
        ]
        ventas = [
            movimiento
            for movimiento in self.movimientos
            if movimiento["estado"] == "aprobada" and movimiento["tipo"] == "venta"
        ]
        canceladas = [
            movimiento
            for movimiento in self.movimientos
            if movimiento["estado"] == "cancelada"
        ]

        self._mostrar_grupo_movimientos(
            compras_aprobadas,
            "COMPRAS APROBADAS",
            "No hay compras aprobadas.",
        )
        self._mostrar_grupo_movimientos(
            ventas,
            "VENTAS",
            "No hay ventas registradas.",
        )
        self._mostrar_grupo_movimientos(
            canceladas,
            "CANCELADAS",
            "No hay movimientos cancelados.",
        )

        filas_resumen = self._filas_resumen_movimientos_por_accion()
        if filas_resumen:
            print("\n--- RESUMEN GENERAL POR ACCION ---")
            self._tabla_markdown(
                filas_resumen,
                columnas_derecha={
                    "Cantidad acciones",
                    "Total inversiones",
                    "Cantidad cancelada",
                    "Total cancelado",
                    "Cantidad vendida",
                    "Precio compra promedio",
                    "Ganancia / perdida",
                },
            )

    def ver_aprobados(self):
        aprobados = [
            movimiento
            for movimiento in self.movimientos
            if movimiento["estado"] == "aprobada"
        ]

        if not aprobados:
            print("No hay movimientos aprobados.")
            return

        self._mostrar_tabla_movimientos(aprobados, "MOVIMIENTOS APROBADOS")

    def resumen_por_accion(self):
        if not self.movimientos:
            print("No hay movimientos registrados.")
            return

        aprobados = [
            movimiento
            for movimiento in self.movimientos
            if movimiento["estado"] == "aprobada"
        ]

        if not aprobados:
            print("No hay movimientos aprobados para resumir.")
            return

        resumen = {}
        for movimiento in aprobados:
            accion = movimiento["nombre_accion"]
            if accion not in resumen:
                resumen[accion] = {
                    "Accion": accion,
                    "Cantidad comprada": 0,
                    "Total comprado": 0.0,
                    "Cantidad vendida": 0,
                    "Total vendido": 0.0,
                }

            if movimiento["tipo"] == "compra":
                resumen[accion]["Cantidad comprada"] += movimiento["cantidad"]
                resumen[accion]["Total comprado"] += movimiento["precio_total"]
            else:
                resumen[accion]["Cantidad vendida"] += movimiento["cantidad"]
                resumen[accion]["Total vendido"] += movimiento["precio_total"]

        filas = []
        for fila in resumen.values():
            fila["Acciones restantes"] = (
                fila["Cantidad comprada"] - fila["Cantidad vendida"]
            )
            fila["Balance dinero"] = fila["Total vendido"] - fila["Total comprado"]
            filas.append(
                {
                    "Acción": fila["Accion"],
                    "Cantidad comprada": self._formatear_numero(
                        fila["Cantidad comprada"]
                    ),
                    "Total comprado": self._formatear_numero(fila["Total comprado"]),
                    "Cantidad vendida": self._formatear_numero(
                        fila["Cantidad vendida"]
                    ),
                    "Total vendido": self._formatear_numero(fila["Total vendido"]),
                    "Acciones restantes": self._formatear_numero(
                        fila["Acciones restantes"]
                    ),
                    "Balance dinero": self._formatear_numero(fila["Balance dinero"]),
                }
            )

        print("\n--- RESUMEN POR ACCION ---")
        self._tabla_markdown(
            filas,
            columnas_derecha={
                "Cantidad comprada",
                "Total comprado",
                "Cantidad vendida",
                "Total vendido",
                "Acciones restantes",
                "Balance dinero",
            },
        )

    def exportar_excel(self, nombre_archivo="movimientos_acciones.xlsx"):
        if not self.movimientos:
            print("No hay movimientos para exportar.")
            return

        filas = self._filas_movimientos(self.movimientos)
        columnas = [
            "Acción",
            "Tipo",
            "Estado",
            "Cantidad",
            "Precio unidad",
            "Comisión",
            "Total",
            "Fecha",
        ]

        if nombre_archivo.lower().endswith(".xlsx"):
            try:
                from openpyxl import Workbook
            except ImportError:
                nombre_csv = nombre_archivo[:-5] + ".csv"
                with open(nombre_csv, "w", newline="", encoding="utf-8-sig") as archivo:
                    escritor = csv.DictWriter(archivo, fieldnames=columnas)
                    escritor.writeheader()
                    escritor.writerows(filas)
                print(
                    f"No se encontro openpyxl. Se exporto un CSV compatible con Excel: {nombre_csv}"
                )
                return

            libro = Workbook()
            hoja = libro.active
            hoja.title = "Movimientos"
            hoja.append(columnas)
            for fila in filas:
                hoja.append([fila[columna] for columna in columnas])
            libro.save(nombre_archivo)
            print(f"Archivo exportado correctamente como: {nombre_archivo}")
            return

        with open(nombre_archivo, "w", newline="", encoding="utf-8-sig") as archivo:
            escritor = csv.DictWriter(archivo, fieldnames=columnas)
            escritor.writeheader()
            escritor.writerows(filas)

        print(f"Archivo exportado correctamente como: {nombre_archivo}")

    def insertar_movimiento(self):
        print("\n--- INGRESAR NUEVO MOVIMIENTO ---")

        nombre_accion = self._pedir_texto(
            "Nombre de la accion: ", convertir_mayusculas=True
        )
        tipo = self._pedir_tipo(
            "Tipo de movimiento (compra/venta) [Enter = compra]: ",
            valor_defecto="compra",
        )
        estado = self._pedir_estado("Estado (aprobada/cancelada): ")
        fecha = self._pedir_fecha(
            "Fecha (YYYY-MM-DD HH:MM, YYYY-MM-DD, MM-DD HH:MM o MM-DD): "
        )
        cantidad = self._pedir_entero("Cantidad de acciones: ")
        precio_unidad = self._pedir_flotante("Precio por unidad: ")
        comision = self._pedir_flotante("Comision (si no hay, escribe 0): ")

        movimiento = self._normalizar_movimiento(
            {
                "nombre_accion": nombre_accion,
                "tipo": tipo,
                "estado": estado,
                "fecha": fecha,
                "cantidad": cantidad,
                "precio_unidad": precio_unidad,
                "comision": comision,
            }
        )
        self.movimientos.append(movimiento)
        print("Movimiento guardado correctamente.\n")

    def modificar_movimiento(self):
        indice = self._seleccionar_movimiento("modificar")
        if indice is None:
            return

        movimiento = self.movimientos[indice].copy()

        print("\nDeja el campo vacio si quieres conservar el valor actual.")

        movimiento["nombre_accion"] = self._pedir_texto(
            f"Nombre de la accion [{movimiento['nombre_accion']}]: ",
            valor_actual=movimiento["nombre_accion"],
            convertir_mayusculas=True,
        )
        movimiento["tipo"] = self._pedir_tipo(
            f"Tipo de movimiento [{movimiento['tipo']}]: ",
            valor_actual=movimiento["tipo"],
        )
        movimiento["estado"] = self._pedir_estado(
            f"Estado [{movimiento['estado']}]: ",
            valor_actual=movimiento["estado"],
        )
        movimiento["fecha"] = self._pedir_fecha(
            f"Fecha [{self._formatear_fecha(movimiento['fecha'])}]: ",
            valor_actual=movimiento["fecha"],
        )
        movimiento["cantidad"] = self._pedir_entero(
            f"Cantidad de acciones [{self._formatear_numero(movimiento['cantidad'])}]: ",
            valor_actual=movimiento["cantidad"],
        )
        movimiento["precio_unidad"] = self._pedir_flotante(
            f"Precio por unidad [{self._formatear_numero(movimiento['precio_unidad'])}]: ",
            valor_actual=movimiento["precio_unidad"],
        )
        movimiento["comision"] = self._pedir_flotante(
            f"Comision [{self._formatear_numero(movimiento['comision'])}]: ",
            valor_actual=movimiento["comision"],
        )

        self.movimientos[indice] = self._normalizar_movimiento(movimiento)
        print("Movimiento modificado correctamente.")

    def _acumular_resumen(self, resumen, movimiento):
        accion = movimiento["nombre_accion"]
        if accion not in resumen:
            resumen[accion] = {
                "Accion": accion,
                "Cantidad comprada": 0,
                "Inversion compras": 0.0,
                "Total comprado": 0.0,
                "Cantidad vendida": 0,
                "Ventas reinvertidas": 0.0,
                "Comisiones": 0.0,
            }

        inversion_bruta = self._calcular_inversion_bruta(
            movimiento["cantidad"],
            movimiento["precio_unidad"],
            movimiento["estado"],
        )
        resumen[accion]["Comisiones"] += movimiento["comision"]

        if movimiento["tipo"] == "compra":
            resumen[accion]["Cantidad comprada"] += movimiento["cantidad"]
            resumen[accion]["Inversion compras"] += inversion_bruta
            resumen[accion]["Total comprado"] += movimiento["precio_total"]
        else:
            resumen[accion]["Cantidad vendida"] += movimiento["cantidad"]
            resumen[accion]["Ventas reinvertidas"] += movimiento["precio_total"]

    def _construir_resumen_numerico_por_accion(self, aprobados):
        resumen = {}
        for movimiento in aprobados:
            self._acumular_resumen(resumen, movimiento)

        filas = []
        for fila in resumen.values():
            cantidad_comprada = fila["Cantidad comprada"]
            cantidad_vendida = fila["Cantidad vendida"]
            acciones_vigentes = cantidad_comprada - cantidad_vendida
            costo_promedio_compra = (
                fila["Total comprado"] / cantidad_comprada
                if cantidad_comprada > 0
                else 0.0
            )
            costo_portafolio_actual = acciones_vigentes * costo_promedio_compra
            resultado_realizado = (
                fila["Ventas reinvertidas"]
                - (cantidad_vendida * costo_promedio_compra)
            )

            filas.append(
                {
                    "Accion": fila["Accion"],
                    "Cantidad comprada": cantidad_comprada,
                    "Total comprado": fila["Total comprado"],
                    "Cantidad vendida": cantidad_vendida,
                    "Ventas netas": fila["Ventas reinvertidas"],
                    "Comisiones": fila["Comisiones"],
                    "Acciones vigentes": acciones_vigentes,
                    "Costo promedio compra": costo_promedio_compra,
                    "Costo portafolio actual": costo_portafolio_actual,
                    "Resultado realizado": resultado_realizado,
                }
            )

        return sorted(
            filas,
            key=lambda fila: (-fila["Costo portafolio actual"], fila["Accion"]),
        )

    def _calcular_totales_generales(self, aprobados):
        total_inversion = 0.0
        total_comisiones = 0.0
        resumen_por_accion = self._construir_resumen_numerico_por_accion(aprobados)

        for movimiento in aprobados:
            total_comisiones += movimiento["comision"]
            if movimiento["tipo"] == "compra":
                total_inversion += self._calcular_inversion_bruta(
                    movimiento["cantidad"],
                    movimiento["precio_unidad"],
                    movimiento["estado"],
                )

        total_compras = sum(fila["Total comprado"] for fila in resumen_por_accion)
        total_ventas = sum(fila["Ventas netas"] for fila in resumen_por_accion)
        acciones_vigentes = sum(
            fila["Acciones vigentes"] for fila in resumen_por_accion
        )
        costo_portafolio_actual = sum(
            fila["Costo portafolio actual"] for fila in resumen_por_accion
        )
        resultado_realizado = sum(
            fila["Resultado realizado"] for fila in resumen_por_accion
        )
        flujo_neto_invertido = total_compras - total_ventas
        valor_real_menos_comisiones = self.VALOR_REAL_HOY - total_comisiones

        return {
            "inversion_compras": total_inversion,
            "comisiones": total_comisiones,
            "total_comprado": total_compras,
            "ventas_netas": total_ventas,
            "flujo_neto_invertido": flujo_neto_invertido,
            "acciones_vigentes": acciones_vigentes,
            "costo_portafolio_actual": costo_portafolio_actual,
            "resultado_realizado": resultado_realizado,
            "valor_real_hoy": self.VALOR_REAL_HOY,
            "valor_real_menos_comisiones": valor_real_menos_comisiones,
        }

    def _fila_totales_generales(self, aprobados):
        totales = self._calcular_totales_generales(aprobados)
        return {
            "Inversion compras": self._formatear_numero(totales["inversion_compras"]),
            "Comisiones": self._formatear_numero(totales["comisiones"]),
            "Total comprado": self._formatear_numero(totales["total_comprado"]),
            "Ventas netas": self._formatear_numero(
                totales["ventas_netas"]
            ),
            "Flujo neto invertido": self._formatear_numero(
                totales["flujo_neto_invertido"]
            ),
            "Acciones vigentes": self._formatear_numero(totales["acciones_vigentes"]),
            "Costo portafolio actual": self._formatear_numero(
                totales["costo_portafolio_actual"]
            ),
            "Resultado realizado": self._formatear_numero(
                totales["resultado_realizado"]
            ),
        }

    def _filas_resumen_financiero(self, aprobados):
        totales = self._calcular_totales_generales(aprobados)
        return [
            {
                "Concepto": "Inversion compras",
                "Valor": self._formatear_numero(totales["inversion_compras"]),
            },
            {
                "Concepto": "Total comprado",
                "Valor": self._formatear_numero(totales["total_comprado"]),
            },
            {
                "Concepto": "Ventas netas",
                "Valor": self._formatear_numero(totales["ventas_netas"]),
            },
            {
                "Concepto": "Flujo neto invertido",
                "Valor": self._formatear_numero(totales["flujo_neto_invertido"]),
            },
            {
                "Concepto": "Costo portafolio actual",
                "Valor": self._formatear_numero(totales["costo_portafolio_actual"]),
            },
            {
                "Concepto": "Resultado realizado",
                "Valor": self._formatear_numero(totales["resultado_realizado"]),
            },
            {
                "Concepto": "Comisiones acumuladas",
                "Valor": self._formatear_numero(totales["comisiones"]),
            },
            {
                "Concepto": "Acciones vigentes",
                "Valor": self._formatear_numero(totales["acciones_vigentes"]),
            },
            {
                "Concepto": "Valor real hoy",
                "Valor": self._formatear_numero(
                    totales["valor_real_hoy"]
                ),
            },
            {
                "Concepto": "Valor real en acciones menos comisiones",
                "Valor": self._formatear_numero(
                    totales["valor_real_menos_comisiones"]
                ),
            },
            {
                "Concepto": "Retiro personal",
                "Valor": self._formatear_numero(self.RETIRO_PERSONAL),
            },
            {
                "Concepto": "Fondos",
                "Valor": self._formatear_numero(self.FONDOS_ACTUALES),
            },
            {
                "Concepto": "Disponible para invertir (dinero a favor)",
                "Valor": self._formatear_numero(self.DISPONIBLE_PARA_INVERTIR),
            },
        ]

    def _obtener_aprobados(self):
        return [
            movimiento
            for movimiento in self.movimientos
            if movimiento["estado"] == "aprobada"
        ]

    def resumen_por_accion(self):
        if not self.movimientos:
            print("No hay movimientos registrados.")
            return

        aprobados = self._obtener_aprobados()

        if not aprobados:
            print("No hay movimientos aprobados para resumir.")
            return

        filas = []
        for fila in self._construir_resumen_numerico_por_accion(aprobados):
            filas.append(
                {
                    "Accion": fila["Accion"],
                    "Cantidad comprada": self._formatear_numero(
                        fila["Cantidad comprada"]
                    ),
                    "Total comprado": self._formatear_numero(
                        fila["Total comprado"]
                    ),
                    "Cantidad vendida": self._formatear_numero(
                        fila["Cantidad vendida"]
                    ),
                    "Ventas netas": self._formatear_numero(
                        fila["Ventas netas"]
                    ),
                    "Comisiones": self._formatear_numero(fila["Comisiones"]),
                    "Acciones vigentes": self._formatear_numero(
                        fila["Acciones vigentes"]
                    ),
                    "Costo promedio compra": self._formatear_numero(
                        fila["Costo promedio compra"]
                    ),
                    "Costo portafolio actual": self._formatear_numero(
                        fila["Costo portafolio actual"]
                    ),
                    "Resultado realizado": self._formatear_numero(
                        fila["Resultado realizado"]
                    ),
                }
            )

        print("\n--- RESUMEN POR ACCION ---")
        self._tabla_markdown(
            filas,
            columnas_derecha={
                "Cantidad comprada",
                "Total comprado",
                "Cantidad vendida",
                "Ventas netas",
                "Comisiones",
                "Acciones vigentes",
                "Costo promedio compra",
                "Costo portafolio actual",
                "Resultado realizado",
            },
        )

        print("\n--- TOTALES GENERALES SIN CANCELADAS ---")
        self._tabla_markdown(
            [self._fila_totales_generales(aprobados)],
            columnas_derecha={
                "Inversion compras",
                "Comisiones",
                "Total comprado",
                "Ventas netas",
                "Flujo neto invertido",
                "Acciones vigentes",
                "Costo portafolio actual",
                "Resultado realizado",
            },
        )

    def ver_resumen_financiero(self):
        if not self.movimientos:
            print("No hay movimientos registrados.")
            return

        aprobados = self._obtener_aprobados()
        if not aprobados:
            print("No hay movimientos aprobados para resumir.")
            return

        print("\n--- RESUMEN FINANCIERO ---")
        self._tabla_markdown(
            self._filas_resumen_financiero(aprobados),
            columnas_derecha={"Valor"},
        )

    def _leer_filas_desde_archivo(self, nombre_archivo):
        if nombre_archivo.lower().endswith(".xlsx"):
            try:
                from openpyxl import load_workbook
            except ImportError as error:
                raise ImportError(
                    "Para importar .xlsx instala openpyxl o usa un archivo .csv."
                ) from error

            libro = load_workbook(nombre_archivo, data_only=True)
            hoja = libro.active
            filas = list(hoja.iter_rows(values_only=True))
            libro.close()
            return filas

        with open(nombre_archivo, newline="", encoding="utf-8-sig") as archivo:
            lector = csv.reader(archivo)
            return list(lector)

    def _mapear_columnas_importacion(self, encabezados):
        mapa = {}
        for indice, encabezado in enumerate(encabezados):
            clave_normalizada = self._normalizar_clave_importacion(encabezado)
            for campo, aliases in self.ALIASES_COLUMNAS_IMPORTACION.items():
                if clave_normalizada in aliases and campo not in mapa:
                    mapa[campo] = indice

        faltantes = [
            campo
            for campo in self.ALIASES_COLUMNAS_IMPORTACION
            if campo not in mapa
        ]
        if faltantes:
            raise ValueError(
                "Faltan columnas requeridas en el archivo: "
                + ", ".join(faltantes)
            )

        return mapa

    def importar_excel(self, nombre_archivo="movimientos_acciones.xlsx"):
        filas = self._leer_filas_desde_archivo(nombre_archivo)
        if not filas:
            print("El archivo no tiene datos.")
            return

        encabezados = filas[0]
        datos = filas[1:]
        mapa = self._mapear_columnas_importacion(encabezados)
        movimientos_importados = []

        for numero_fila, fila in enumerate(datos, start=2):
            valores = list(fila)
            while len(valores) < len(encabezados):
                valores.append(None)

            if all(
                valor is None or str(valor).strip() == ""
                for valor in valores[: len(encabezados)]
            ):
                continue

            try:
                movimiento = self._normalizar_movimiento(
                    {
                        "nombre_accion": valores[mapa["nombre_accion"]],
                        "tipo": valores[mapa["tipo"]],
                        "estado": valores[mapa["estado"]],
                        "fecha": valores[mapa["fecha"]],
                        "cantidad": self._parsear_numero(
                            str(valores[mapa["cantidad"]])
                        ),
                        "precio_unidad": self._parsear_numero(
                            str(valores[mapa["precio_unidad"]])
                        ),
                        "comision": self._parsear_numero(
                            str(valores[mapa["comision"]] or 0)
                        ),
                    }
                )
            except Exception as error:
                print(f"No se pudo importar la fila {numero_fila}: {error}")
                return

            movimientos_importados.append(movimiento)

        self.movimientos = movimientos_importados
        print(
            f"Se importaron {len(self.movimientos)} movimientos desde {nombre_archivo}."
        )

    def pedir_importacion(self):
        nombre_archivo = (
            input("Nombre del archivo a importar [.xlsx o .csv]: ").strip()
            or "movimientos_acciones.xlsx"
        )
        self.importar_excel(nombre_archivo)

    def exportar_excel(self, nombre_archivo="movimientos_acciones.xlsx"):
        if not self.movimientos:
            print("No hay movimientos para exportar.")
            return

        filas = self._filas_movimientos(self.movimientos)
        columnas = [
            "Accion",
            "Tipo",
            "Estado",
            "Cantidad",
            "Precio unidad",
            "Inversion",
            "Comision",
            "Total",
            "Fecha",
        ]

        if nombre_archivo.lower().endswith(".xlsx"):
            try:
                from openpyxl import Workbook
            except ImportError:
                nombre_csv = nombre_archivo[:-5] + ".csv"
                with open(nombre_csv, "w", newline="", encoding="utf-8-sig") as archivo:
                    escritor = csv.DictWriter(archivo, fieldnames=columnas)
                    escritor.writeheader()
                    escritor.writerows(filas)
                print(
                    f"No se encontro openpyxl. Se exporto un CSV compatible con Excel: {nombre_csv}"
                )
                return

            libro = Workbook()
            hoja = libro.active
            hoja.title = "Movimientos"
            hoja.append(columnas)
            for fila in filas:
                hoja.append([fila[columna] for columna in columnas])
            libro.save(nombre_archivo)
            print(f"Archivo exportado correctamente como: {nombre_archivo}")
            return

        with open(nombre_archivo, "w", newline="", encoding="utf-8-sig") as archivo:
            escritor = csv.DictWriter(archivo, fieldnames=columnas)
            escritor.writeheader()
            escritor.writerows(filas)

        print(f"Archivo exportado correctamente como: {nombre_archivo}")


def obtener_comando_dashboard():
    directorio_base = Path(__file__).resolve().parent
    archivo_app = directorio_base / "streamlit_app.py"

    if not archivo_app.exists():
        raise FileNotFoundError("No se encontro streamlit_app.py.")

    candidatos = [
        Path(sys.executable).resolve().parent / "streamlit.exe",
        directorio_base / "venv" / "Scripts" / "streamlit.exe",
    ]

    for ejecutable in candidatos:
        if ejecutable.exists():
            return [str(ejecutable), "run", str(archivo_app)], directorio_base

    return [sys.executable, "-m", "streamlit", "run", str(archivo_app)], directorio_base


def ver_dashboard():
    try:
        comando, directorio_base = obtener_comando_dashboard()
    except Exception as error:
        print(f"No fue posible preparar el dashboard: {error}")
        return

    try:
        subprocess.Popen(
            comando,
            cwd=str(directorio_base),
            creationflags=getattr(subprocess, "CREATE_NEW_CONSOLE", 0),
        )
    except Exception as error:
        print(f"No fue posible iniciar el dashboard: {error}")
        return

    print("Dashboard iniciado en una nueva ventana.")
    print("Si no se abre automaticamente, entra a http://localhost:8501")


def menu():
    fondo = FondoAcciones()

    while True:
        print("\n===== MENU =====")
        print("1. Insertar movimiento")
        print("2. Importar desde Excel/CSV")
        print("3. Ver todos los movimientos")
        print("4. Ver movimientos aprobados")
        print("5. Ver resumen por accion")
        print("6. Ver resumen financiero")
        print("7. Exportar a Excel")
        print("8. Modificar movimiento")
        print("9. Eliminar movimiento")
        print("10. Ver dashboard")
        print("11. Salir")

        opcion = input("Elige una opcion: ").strip()

        if opcion == "1":
            fondo.insertar_movimiento()
        elif opcion == "2":
            fondo.pedir_importacion()
        elif opcion == "3":
            fondo.ver_movimientos()
        elif opcion == "4":
            fondo.ver_aprobados()
        elif opcion == "5":
            fondo.resumen_por_accion()
        elif opcion == "6":
            fondo.ver_resumen_financiero()
        elif opcion == "7":
            fondo.exportar_excel()
        elif opcion == "8":
            fondo.modificar_movimiento()
        elif opcion == "9":
            fondo.eliminar_movimiento()
        elif opcion == "10":
            ver_dashboard()
        elif opcion == "11":
            print("Saliendo del programa...")
            break
        else:
            print("Opcion no valida.")


if __name__ == "__main__":
    menu()
